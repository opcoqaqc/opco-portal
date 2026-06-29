"""Extract QA/QC team-member KPI data from an
OPCO_QAQC_<Member>_KPI.xlsx workbook (5 sheets: Cover / KPI
Definitions / Scoring Matrix / Monthly Tracker / Data Sources).

The team-member workbooks use the TKPI scheme (with role-split
weights), so the output JSON is shaped slightly differently from
the manager: TKPI-N ids, engineer + inspector measurements
flattened into a single description, and the role-appropriate
weight chosen based on the Cover sheet's "Role" field.

Run:  py scripts/extract_qaqc_member.py "<xlsx-path>" > <out>.json
"""
import json
import sys
from datetime import datetime

import openpyxl

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def cell(row, idx):
    if row is None or idx >= len(row):
        return None
    return row[idx]


def s(v):
    if v is None:
        return ''
    return str(v).strip()


def num(v):
    if v is None or v == '' or v == '—':
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# --- Cover --------------------------------------------------------------
def extract_cover(sh):
    rows = list(sh.iter_rows(values_only=True))
    meta = {}
    keys = {
        'Managed By':         'managed_by',
        'Team Member':        'member',
        'Role':               'role',
        'Reporting Period':   'reporting_period',
        'Document Reference': 'doc_ref',
        'Revision':           'revision',
        'Submission Date':    'submission_date',
    }
    title = ''
    for row in rows:
        v = s(cell(row, 1))
        if v.startswith('Individual Performance'):
            title = v
            break

    for row in rows:
        label = s(cell(row, 1))
        if label in keys:
            meta[keys[label]] = s(cell(row, 2))

    return {
        'title':     title or 'Team-member Performance',
        'company':   'ORIENT PIPELINE CO.',
        'meta':      meta,
        'framework': [],
        'standards': '',
    }


def detect_role(cover_meta):
    """Return 'engineer' or 'inspector' based on Cover.Role."""
    r = (cover_meta or {}).get('role', '').lower()
    if 'engineer' in r:
        return 'engineer'
    return 'inspector'


# --- KPI Definitions ---------------------------------------------------
def extract_definitions(sh, role):
    rows = list(sh.iter_rows(values_only=True))
    items = []
    for row in rows:
        tid = s(cell(row, 1))
        if not tid.startswith('TKPI-'):
            continue
        name        = s(cell(row, 2))
        purpose     = s(cell(row, 3))
        eng_metric  = s(cell(row, 4))
        ins_metric  = s(cell(row, 5))
        unit        = s(cell(row, 6))
        target      = s(cell(row, 7))

        role_metric = eng_metric if role == 'engineer' else ins_metric
        other_metric = ins_metric if role == 'engineer' else eng_metric
        other_label = 'Inspector' if role == 'engineer' else 'Engineer'
        role_label  = 'Engineer'  if role == 'engineer' else 'Inspector'

        # Stack purpose + this role's measurement up top, then the
        # other role's measurement as a context note. Keeps the
        # description block focused on the member's own job.
        description = purpose.strip()
        methodology = f'{role_label} measurement: {role_metric}'
        if other_metric:
            methodology += f'\n\n{other_label} measurement (for reference): {other_metric}'

        items.append({
            'id':           tid,
            'name':         name,
            'description':  description,
            'methodology':  methodology,
            'unit':         unit,
            'target':       target,
            'data_source':  '',  # filled in by extract_sources()
        })
    return items


# --- Scoring Matrix ----------------------------------------------------
def extract_scoring(sh, role):
    rows = list(sh.iter_rows(values_only=True))
    kpis = []
    for row in rows:
        tid = s(cell(row, 1))
        if not tid.startswith('TKPI-'):
            continue
        name             = s(cell(row, 2))
        eng_weight       = num(cell(row, 3))
        ins_weight       = num(cell(row, 4))
        bands            = [s(cell(row, i)) for i in range(5, 10)]
        weight = eng_weight if role == 'engineer' else ins_weight
        kpis.append({
            'id':     tid,
            'name':   name,
            'weight': weight,
            'bands':  bands,
        })

    composite = []
    in_block = False
    for row in rows:
        v = s(cell(row, 1))
        if v == 'COMPOSITE PERFORMANCE RATING':
            in_block = True
            continue
        if in_block and v and not v.startswith('Composite Score'):
            composite.append({
                'range':          v,
                'rating':         s(cell(row, 3)),
                'interpretation': s(cell(row, 5)),
            })
    return {'kpis': kpis, 'composite': composite}


# --- Monthly Tracker ---------------------------------------------------
def extract_tracker(sh):
    rows = list(sh.iter_rows(values_only=True))

    def month_vals(row):
        return [num(cell(row, i)) for i in range(3, 15)]

    metrics = []
    composite_rows = []
    in_composite = False
    for row in rows:
        label = s(cell(row, 1))
        if not label:
            continue
        if label == 'Metric':
            continue
        if label.startswith('COMPOSITE PERFORMANCE'):
            in_composite = True
            continue
        if label == 'PERFORMANCE RATING' or label.startswith('COMPOSITE SCORE') \
                or label.startswith('Yellow-shaded'):
            in_composite = False
            continue
        if in_composite:
            composite_rows.append({
                'kpi':      label,
                'expr':     s(cell(row, 2)),
                'weight':   num(cell(row, 15)),
                'score':    num(cell(row, 16)),
                'weighted': num(cell(row, 17)),
            })
            continue

        unit = s(cell(row, 2))
        if not unit:
            metrics.append({'label': label, 'role': 'section'})
            continue

        metrics.append({
            'label':  label,
            'unit':   unit,
            'months': month_vals(row),
            'ytd':    num(cell(row, 15)),
            'target': s(cell(row, 16)),
            'score':  num(cell(row, 17)),
            'role':   'metric',
        })

    rating_text = ''
    composite_score = None
    for row in rows:
        if s(cell(row, 1)) == 'PERFORMANCE RATING':
            rating_text = s(cell(row, 17))
        if s(cell(row, 1)).startswith('COMPOSITE SCORE'):
            composite_score = num(cell(row, 17))

    return {
        'year':            2026,
        'months':          MONTHS,
        'metrics':         metrics,
        'composite_rows':  composite_rows,
        'composite_score': composite_score,
        'rating':          rating_text,
    }


# --- Data Sources ------------------------------------------------------
def extract_sources(sh):
    """Map TKPI-style data-sources columns to manager-shape rows.

    Member workbook columns:
        TKPI #, KPI Name, Input Metric(s), Source Document / System,
        Data Owner, Frequency, Verification / Cross-Check
    """
    rows = list(sh.iter_rows(values_only=True))
    items = []
    governance = []
    in_gov = False
    for row in rows:
        tid = s(cell(row, 1))
        if not tid:
            # Trailing "Data Integrity:" notes after the table become governance
            text = s(cell(row, 2))
            if text and ':' in text:
                label, _, body = text.partition(':')
                governance.append({'label': label.strip(), 'text': body.strip()})
            continue
        if tid == 'GOVERNANCE & CONTROL PRINCIPLES':
            in_gov = True
            continue
        if in_gov:
            label = s(cell(row, 1))
            txt   = s(cell(row, 3))
            if label and txt:
                governance.append({'label': label, 'text': txt})
            continue
        if not tid.startswith('TKPI-'):
            continue
        items.append({
            'id':         tid,
            'kpi':        s(cell(row, 2)),
            'source':     s(cell(row, 3)),
            'system':     s(cell(row, 4)),
            'owner':      s(cell(row, 5)),
            'frequency':  s(cell(row, 6)),
            'retention':  s(cell(row, 7)),
        })
    return {'rows': items, 'governance': governance}


def main():
    if len(sys.argv) < 2:
        print('usage: py extract_qaqc_member.py <xlsx-path>', file=sys.stderr)
        sys.exit(2)
    xlsx = sys.argv[1]
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)

    cover = extract_cover(wb['Cover'])
    role  = detect_role(cover['meta'])

    defs   = extract_definitions(wb['KPI Definitions'], role)
    scoring = extract_scoring(wb['Scoring Matrix'], role)
    tracker = extract_tracker(wb['Monthly Tracker'])
    sources = extract_sources(wb['Data Sources'])

    # Backfill data_source on each definition from the data_sources rows
    src_by_id = {r['id']: (r.get('source','') + ' · ' + r.get('system','')).strip(' ·')
                 for r in sources['rows']}
    for d in defs:
        d['data_source'] = src_by_id.get(d['id'], '')

    out = {
        'data_date':       datetime.now().strftime("%Y-%m-%d"),
        'source_file':     xlsx.split('\\')[-1],
        'cover':           cover,
        'kpi_definitions': defs,
        'scoring_matrix':  scoring,
        'monthly_tracker': tracker,
        'data_sources':    sources,
    }
    sys.stdout.reconfigure(encoding='utf-8')
    print(json.dumps(out, ensure_ascii=False, default=str))


if __name__ == '__main__':
    main()
