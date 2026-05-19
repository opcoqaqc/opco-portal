"""Extract QA/QC Manager KPI data from the OPCO_QAQC_Manager_KPI_*.xlsx
workbook. Five sheets become five JSON sections — Cover, KPI Definitions,
Scoring Matrix, Monthly Tracker, Data Sources.

Run:  py scripts/extract_qaqc.py > .qaqc.json
"""
import json
import sys
from datetime import datetime

import openpyxl

XLSX = r"C:\Users\Msi\Downloads\OPCO_QAQC_Manager_KPI_7.xlsx"

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def cell(row, idx):
    """Safe row[idx] -> returns None if out of range."""
    if row is None or idx >= len(row):
        return None
    return row[idx]


def s(v):
    if v is None:
        return ''
    return str(v).strip()


def num(v):
    if v is None or v == '' or v == '—':
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# --- Cover --------------------------------------------------------------
def extract_cover(sh):
    rows = list(sh.iter_rows(values_only=True))
    # Header rows: 9-17 are label : value pairs (col B + col C, 1-indexed -> idx 1,2)
    meta = {}
    keys = {
        'Position': 'position',
        'Department': 'department',
        'Reporting To': 'reporting_to',
        'Scope of Authority': 'scope',
        'Reporting Period': 'reporting_period',
        'Review Cadence': 'review_cadence',
        'Document Reference': 'doc_ref',
        'Revision': 'revision',
        'Submission Date': 'submission_date',
    }
    for row in rows:
        label = s(cell(row, 1))
        if label in keys:
            meta[keys[label]] = s(cell(row, 2))

    # Framework: rows 21-24 are "1. KPI Definitions" -> description
    framework = []
    for row in rows:
        label = s(cell(row, 1))
        if label and label[0].isdigit() and '. ' in label:
            framework.append({'name': label, 'desc': s(cell(row, 2))})

    standards = ''
    for row in rows:
        v = s(cell(row, 1))
        if v.startswith('Aligned with:'):
            standards = v.replace('Aligned with:', '').strip()
            break

    return {
        'title': 'Key Performance Indicators — QA/QC Management',
        'company': 'ORIENT PIPELINE CO.',
        'meta': meta,
        'framework': framework,
        'standards': standards,
    }


# --- KPI Definitions ---------------------------------------------------
def extract_definitions(sh):
    rows = list(sh.iter_rows(values_only=True))
    items = []
    for row in rows:
        kpi_id = s(cell(row, 1))
        if not kpi_id.startswith('KPI-'):
            continue
        items.append({
            'id':           kpi_id,
            'name':         s(cell(row, 2)),
            'description':  s(cell(row, 3)),
            'methodology':  s(cell(row, 4)),
            'unit':         s(cell(row, 5)),
            'target':       s(cell(row, 6)),
            'data_source':  s(cell(row, 7)),
        })
    return items


# --- Scoring Matrix ----------------------------------------------------
def extract_scoring(sh):
    rows = list(sh.iter_rows(values_only=True))
    kpis = []
    for row in rows:
        kpi_id = s(cell(row, 1))
        if not kpi_id.startswith('KPI-'):
            continue
        kpis.append({
            'id':     kpi_id,
            'name':   s(cell(row, 2)),
            'weight': num(cell(row, 3)),
            'bands':  [s(cell(row, i)) for i in range(4, 9)],   # 5 bands
        })
    # Composite rating bands (rows 15-19)
    composite = []
    in_block = False
    for row in rows:
        v = s(cell(row, 1))
        if v == 'COMPOSITE PERFORMANCE RATING':
            in_block = True
            continue
        if in_block and v and v != 'Composite Score (Weighted)':
            composite.append({
                'range':          v,
                'rating':         s(cell(row, 3)),
                'interpretation': s(cell(row, 5)),
            })
    return {'kpis': kpis, 'composite': composite}


# --- Monthly Tracker ---------------------------------------------------
def extract_tracker(sh):
    rows = list(sh.iter_rows(values_only=True))
    # Column layout (0-indexed):
    #   1  = metric label (col B)
    #   2  = unit (col C)
    #   3..14 = Jan..Dec (cols D..O)
    #   15 = YTD (col P)
    #   16 = target text (col Q)
    #   17 = score (col R)  — or weight for composite rows
    def month_vals(row):
        return [num(cell(row, i)) for i in range(3, 15)]

    metrics = []
    composite_rows = []
    in_composite = False
    for row in rows:
        label = s(cell(row, 1))
        if not label:
            continue
        if label == 'KPI / Metric':          # column-header row, skip
            continue
        if label == 'COMPOSITE PERFORMANCE — WEIGHTED YTD':
            in_composite = True
            continue
        if label == 'PERFORMANCE RATING' or label.startswith('Yellow-shaded'):
            in_composite = False
            continue
        if label.startswith('COMPOSITE SCORE'):
            in_composite = False
            continue
        if in_composite:
            # Columns for composite contribution: 15=weight, 16=score, 17=weighted
            composite_rows.append({
                'kpi':      label,
                'expr':     s(cell(row, 2)),
                'weight':   num(cell(row, 15)),
                'score':    num(cell(row, 16)),
                'weighted': num(cell(row, 17)),
            })
            continue

        # KPI section header rows (no unit, just a title)
        unit = s(cell(row, 2))
        if not unit and label.startswith('KPI-'):
            metrics.append({'label': label, 'role': 'section'})
            continue

        # Data row
        if unit:
            metrics.append({
                'label':  label,
                'unit':   unit,
                'months': month_vals(row),
                'ytd':    num(cell(row, 15)),
                'target': s(cell(row, 16)),
                'score':  num(cell(row, 17)),
                'role':   'metric',
            })

    # Hero: composite score + rating live at the bottom (col R = idx 17)
    rating_text = ''
    composite_score = None
    for row in rows:
        if s(cell(row, 1)) == 'PERFORMANCE RATING':
            rating_text = s(cell(row, 17))
        if s(cell(row, 1)).startswith('COMPOSITE SCORE'):
            composite_score = num(cell(row, 17))

    return {
        'year':            2026,
        'months':          MONTHS,
        'metrics':         metrics,
        'composite_rows':  composite_rows,
        'composite_score': composite_score,
        'rating':          rating_text,
    }


# --- Data Sources ------------------------------------------------------
def extract_sources(sh):
    rows = list(sh.iter_rows(values_only=True))
    items = []
    governance = []
    in_gov = False
    for row in rows:
        kpi_id = s(cell(row, 1))
        if kpi_id == 'GOVERNANCE & CONTROL PRINCIPLES':
            in_gov = True
            continue
        if in_gov:
            label = s(cell(row, 1))
            text  = s(cell(row, 3))
            if label and text:
                governance.append({'label': label, 'text': text})
            continue
        if kpi_id.startswith('KPI-'):
            # Display override: ISO Audit cadence is yearly, not monthly
            # (spreadsheet still says "Monthly (internal)" for legacy reasons).
            freq = s(cell(row, 6))
            freq = freq.replace('Monthly (internal)', 'Yearly (internal)')
            items.append({
                'id':         kpi_id,
                'kpi':        s(cell(row, 2)),
                'source':     s(cell(row, 3)),
                'system':     s(cell(row, 4)),
                'owner':      s(cell(row, 5)),
                'frequency':  freq,
                'retention':  s(cell(row, 7)),
            })
    return {'rows': items, 'governance': governance}


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    out = {
        'data_date':       datetime.now().strftime("%Y-%m-%d"),
        'source_file':     XLSX.split('\\')[-1],
        'cover':           extract_cover(wb['Cover']),
        'kpi_definitions': extract_definitions(wb['KPI Definitions']),
        'scoring_matrix':  extract_scoring(wb['Scoring Matrix']),
        'monthly_tracker': extract_tracker(wb['Monthly Tracker']),
        'data_sources':    extract_sources(wb['Data Sources']),
    }
    sys.stdout.reconfigure(encoding='utf-8')
    print(json.dumps(out, ensure_ascii=False, default=str))


if __name__ == '__main__':
    main()
