"""Extract QA/QC Team KPI data from the OPCO_QAQC_Team_KPI_*.xlsx
workbook. Eight sheets become five JSON sections:
- cover (header / company / framework / cascade note)
- kpi_definitions (5 TKPIs with engineer + inspector measurements)
- scoring_matrix (TKPI bands + role-split weights + composite rating)
- members (4 per-member monthly trackers + composite)
- team_dashboard (year-end consolidated table)

Run:  py scripts/extract_qaqc_team.py > .qaqc_team.json
"""
import json
import sys
from datetime import datetime

import openpyxl

XLSX = r"C:\Users\Msi\Downloads\OPCO_QAQC_Team_KPI_4.xlsx"

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def cell(row, idx):
    if row is None or idx >= len(row):
        return None
    return row[idx]


def s(v):
    if v is None:
        return ''
    return str(v).strip()


def num(v):
    if v is None or v == '' or v == '—':
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# --- Cover --------------------------------------------------------------
def extract_cover(sh):
    rows = list(sh.iter_rows(values_only=True))
    meta = {}
    keys = {
        'Managed By': 'managed_by',
        'Team Composition': 'team',
        'Purpose': 'purpose',
        'Reporting Period': 'reporting_period',
        'Document Reference': 'doc_ref',
        'Revision': 'revision',
        'Submission Date': 'submission_date',
    }
    for row in rows:
        label = s(cell(row, 1))
        if label in keys:
            meta[keys[label]] = s(cell(row, 2))

    framework = []
    for row in rows:
        label = s(cell(row, 1))
        if label and label[0].isdigit() and '. ' in label:
            framework.append({'name': label, 'desc': s(cell(row, 2))})

    cascade = ''
    for row in rows:
        v = s(cell(row, 1))
        if v.startswith('Cascade Principle'):
            cascade = v
            break

    return {
        'title': 'Team Performance — QA/QC Engineer & Inspectors',
        'company': 'ORIENT PIPELINE CO.',
        'meta': meta,
        'framework': framework,
        'cascade_note': cascade,
    }


# --- KPI Definitions ---------------------------------------------------
def extract_definitions(sh):
    rows = list(sh.iter_rows(values_only=True))
    items = []
    for row in rows:
        tid = s(cell(row, 1))
        if not tid.startswith('TKPI-'):
            continue
        items.append({
            'id':                tid,
            'name':              s(cell(row, 2)),
            'purpose':           s(cell(row, 3)),
            'engineer_metric':   s(cell(row, 4)),
            'inspector_metric':  s(cell(row, 5)),
            'unit':              s(cell(row, 6)),
            'target':            s(cell(row, 7)),
        })
    return items


# --- Scoring Matrix ----------------------------------------------------
def extract_scoring(sh):
    rows = list(sh.iter_rows(values_only=True))
    kpis = []
    for row in rows:
        tid = s(cell(row, 1))
        if not tid.startswith('TKPI-'):
            continue
        kpis.append({
            'id':                tid,
            'name':              s(cell(row, 2)),
            'weight_engineer':   num(cell(row, 3)),
            'weight_inspector':  num(cell(row, 4)),
            'bands':             [s(cell(row, i)) for i in range(5, 10)],
        })

    composite = []
    in_block = False
    for row in rows:
        v = s(cell(row, 1))
        if v == 'COMPOSITE PERFORMANCE RATING':
            in_block = True
            continue
        if in_block and v and v != 'Composite Score':
            composite.append({
                'range':          v,
                'rating':         s(cell(row, 3)),
                'interpretation': s(cell(row, 5)),
            })
    return {'kpis': kpis, 'composite': composite}


# --- Per-Member Monthly Tracker ---------------------------------------
def extract_member(sh):
    """Per-member sheet -> monthly tracker + composite."""
    rows = list(sh.iter_rows(values_only=True))

    # Header row 1 has "Name — Role"
    header = s(cell(rows[0], 1)) if rows else ''
    # Split on em-dash or hyphen (handle different whitespace)
    name, role = header, ''
    for sep in (' — ', ' -- ', ' - '):
        if sep in header:
            parts = header.split(sep, 1)
            name = parts[0].strip()
            role = parts[1].strip() if len(parts) > 1 else ''
            break
    # Fallback: try double-dash variant ("Ahmed Dlshad  —  QC Engineer")
    if not role:
        import re
        m = re.match(r'^(.*?)\s+[—–-]+\s+(.*)$', header)
        if m:
            name, role = m.group(1).strip(), m.group(2).strip()

    # Check for new-hire warning in r3 area
    is_new_hire = False
    new_hire_note = ''
    for row in rows[:5]:
        v = s(cell(row, 1))
        if 'NEW HIRE' in v.upper() or 'new hire' in v.lower():
            is_new_hire = True
            new_hire_note = v
            break

    # Walk metrics — column layout matches Manager tracker (idx 3-14 = Jan-Dec)
    def month_vals(row):
        return [num(cell(row, i)) for i in range(3, 15)]

    metrics = []
    composite_rows = []
    in_composite = False
    for row in rows:
        label = s(cell(row, 1))
        if not label:
            continue
        if label == 'Metric':         # column-header row
            continue
        if label.startswith('COMPOSITE PERFORMANCE'):
            in_composite = True
            continue
        if label == 'PERFORMANCE RATING' or label.startswith('COMPOSITE SCORE'):
            in_composite = False
            continue
        if in_composite:
            composite_rows.append({
                'kpi':      label,
                'expr':     s(cell(row, 2)),
                'weight':   num(cell(row, 15)),
                'score':    num(cell(row, 16)),
                'weighted': num(cell(row, 17)),
            })
            continue

        # Section header rows (TKPI-N · ..., no unit)
        unit = s(cell(row, 2))
        if not unit and label.startswith('TKPI-'):
            metrics.append({'label': label, 'role': 'section'})
            continue

        # Data row
        if unit:
            metrics.append({
                'label':  label,
                'unit':   unit,
                'months': month_vals(row),
                'ytd':    num(cell(row, 15)),
                'target': s(cell(row, 16)),
                'score':  num(cell(row, 17)),
                'role':   'metric',
            })

    # Composite score / rating live at the bottom (col R = idx 17)
    composite_score = None
    rating_text = ''
    for row in rows:
        if s(cell(row, 1)).startswith('COMPOSITE SCORE'):
            composite_score = num(cell(row, 17))
        if s(cell(row, 1)) == 'PERFORMANCE RATING':
            rating_text = s(cell(row, 17))

    return {
        'name':            name,
        'role':            role,
        'is_new_hire':     is_new_hire,
        'new_hire_note':   new_hire_note,
        'metrics':         metrics,
        'composite_rows':  composite_rows,
        'composite_score': composite_score,
        'rating':          rating_text,
        'year':            2026,
        'months':          MONTHS,
    }


# --- Team Dashboard ----------------------------------------------------
def extract_dashboard(sh):
    rows = list(sh.iter_rows(values_only=True))
    members = []
    team_avg = None
    for row in rows:
        label = s(cell(row, 1))
        if not label:
            continue
        if label in ('Team Member',):       # column header
            continue
        if label == 'TEAM AVERAGE':
            team_avg = {
                'scores': [num(cell(row, i)) for i in range(3, 8)],  # 5 TKPI scores
                'composite': num(cell(row, 8)),
            }
            continue
        # New-hire row has multiline name "Ahmed Taha\n(New Hire — May 2026)"
        role = s(cell(row, 2))
        if role not in ('QC Engineer', 'QC Inspector'):
            continue

        # Trim newline-suffixed names
        clean_name = label.split('\n')[0].strip()
        is_new_hire = 'New Hire' in label or 'NEW HIRE' in label.upper()
        members.append({
            'name':       clean_name,
            'name_full':  label,
            'role':       role,
            'is_new_hire': is_new_hire,
            'scores':     [num(cell(row, i)) for i in range(3, 8)],  # 5 TKPI scores
            'composite':  num(cell(row, 8)),
            'rating':     s(cell(row, 9)),
        })

    return {
        'members': members,
        'team_average': team_avg,
    }


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    member_sheets = [n for n in wb.sheetnames
                     if n not in ('Cover', 'KPI Definitions',
                                  'Scoring Matrix', 'Team Dashboard')]
    out = {
        'data_date':       datetime.now().strftime("%Y-%m-%d"),
        'source_file':     XLSX.split('\\')[-1],
        'cover':           extract_cover(wb['Cover']),
        'kpi_definitions': extract_definitions(wb['KPI Definitions']),
        'scoring_matrix':  extract_scoring(wb['Scoring Matrix']),
        'members':         [extract_member(wb[n]) for n in member_sheets],
        'team_dashboard':  extract_dashboard(wb['Team Dashboard']),
    }
    sys.stdout.reconfigure(encoding='utf-8')
    print(json.dumps(out, ensure_ascii=False, default=str))


if __name__ == '__main__':
    main()
