"""Extract HSE KPI data from an OPCO HSE KPI Excel workbook.
Same 5-sheet shape as the QA/QC Manager workbook (Cover / KPI
Definitions / Scoring Matrix / Monthly Tracker / Data Sources),
adapted to HSE-specific quirks:
  - Scoring matrix has an extra "Direction" column (Lower is better / Higher is better)
  - Tracker sections contain INCIDENTS / EXPOSURE HOURS sub-headers
  - Title comes from the Cover sheet (not hardcoded)

Run:  py scripts/extract_hse.py "<xlsx-path>" > <out>.json
"""
import json
import sys
from datetime import datetime

import openpyxl

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def cell(row, idx):
    if row is None or idx >= len(row):
        return None
    return row[idx]


def s(v):
    if v is None:
        return ''
    return str(v).strip()


def num(v):
    if v is None or v == '' or v == '—':
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# --- Cover --------------------------------------------------------------
def extract_cover(sh):
    rows = list(sh.iter_rows(values_only=True))
    meta = {}
    keys = {
        'Position':            'position',
        'Department':          'department',
        'Reporting To':        'reporting_to',
        'Scope of Authority':  'scope',
        'Reporting Period':    'reporting_period',
        'Review Frequency':    'review_cadence',
        'Review Cadence':      'review_cadence',
        'Document Reference':  'doc_ref',
        'Revision':            'revision',
        'Submission Date':     'submission_date',
    }
    title = ''
    for row in rows:
        v = s(cell(row, 1))
        if v.startswith('Key Performance Indicators'):
            title = v
            break

    for row in rows:
        label = s(cell(row, 1))
        if label in keys:
            meta[keys[label]] = s(cell(row, 2))

    framework = []
    for row in rows:
        label = s(cell(row, 1))
        if label and label[0].isdigit() and '. ' in label:
            framework.append({'name': label, 'desc': s(cell(row, 2))})

    standards = ''
    for row in rows:
        v = s(cell(row, 1))
        if v.startswith('Aligned with:'):
            standards = v.replace('Aligned with:', '').strip()
            # Trim the same trailing list we trim in QA/QC, if present
            for tail in (' · ASME B31.3 · ASME Section IX · API 1104 · API 5L · Company QMS',
                         '· ASME B31.3 · ASME Section IX · API 1104 · API 5L · Company QMS'):
                if standards.endswith(tail):
                    standards = standards[:-len(tail)].rstrip(' ·')
                    break
            break

    return {
        'title':     title or 'HSE Key Performance Indicators',
        'company':   'ORIENT PIPELINE CO.',
        'meta':      meta,
        'framework': framework,
        'standards': standards,
    }


# --- KPI Definitions ---------------------------------------------------
def extract_definitions(sh):
    rows = list(sh.iter_rows(values_only=True))
    items = []
    for row in rows:
        kpi_id = s(cell(row, 1))
        if not kpi_id.startswith('KPI-'):
            continue
        items.append({
            'id':           kpi_id,
            'name':         s(cell(row, 2)),
            'description':  s(cell(row, 3)),
            'methodology':  s(cell(row, 4)),
            'unit':         s(cell(row, 5)),
            'target':       s(cell(row, 6)),
            'data_source':  s(cell(row, 7)),
        })
    return items


# --- Scoring Matrix ----------------------------------------------------
def extract_scoring(sh):
    rows = list(sh.iter_rows(values_only=True))
    kpis = []
    for row in rows:
        kpi_id = s(cell(row, 1))
        if not kpi_id.startswith('KPI-'):
            continue
        kpis.append({
            'id':        kpi_id,
            'name':      s(cell(row, 2)),
            'weight':    num(cell(row, 3)),
            'bands':     [s(cell(row, i)) for i in range(4, 9)],   # 5 score bands
            'direction': s(cell(row, 9)),                          # extra HSE column
        })

    composite = []
    in_block = False
    for row in rows:
        v = s(cell(row, 1))
        if v == 'COMPOSITE PERFORMANCE RATING':
            in_block = True
            continue
        if in_block and v and not v.startswith('Composite Score'):
            composite.append({
                'range':          v,
                'rating':         s(cell(row, 3)),
                'interpretation': s(cell(row, 5)),
            })
    return {'kpis': kpis, 'composite': composite}


# --- Monthly Tracker ---------------------------------------------------
def extract_tracker(sh):
    rows = list(sh.iter_rows(values_only=True))

    def month_vals(row):
        return [num(cell(row, i)) for i in range(3, 15)]

    metrics = []
    composite_rows = []
    in_composite = False
    for row in rows:
        label = s(cell(row, 1))
        if not label:
            continue
        if label == 'KPI / Metric':
            continue
        if label == 'COMPOSITE PERFORMANCE — WEIGHTED YTD':
            in_composite = True
            continue
        if label == 'PERFORMANCE RATING' or label.startswith('Yellow-shaded'):
            in_composite = False
            continue
        if label.startswith('COMPOSITE SCORE'):
            in_composite = False
            continue
        if in_composite:
            composite_rows.append({
                'kpi':      label,
                'expr':     s(cell(row, 2)),
                'weight':   num(cell(row, 15)),
                'score':    num(cell(row, 16)),
                'weighted': num(cell(row, 17)),
            })
            continue

        # Section header rows (KPI-N · ... with no unit) and sub-header rows
        # (INCIDENTS, EXPOSURE HOURS, etc.) — both treated as sections so
        # the renderer can group them visually.
        unit = s(cell(row, 2))
        if not unit:
            metrics.append({'label': label, 'role': 'section'})
            continue

        # Data row
        metrics.append({
            'label':  label,
            'unit':   unit,
            'months': month_vals(row),
            'ytd':    num(cell(row, 15)),
            'target': s(cell(row, 16)),
            'score':  num(cell(row, 17)),
            'role':   'metric',
        })

    rating_text = ''
    composite_score = None
    for row in rows:
        if s(cell(row, 1)) == 'PERFORMANCE RATING':
            rating_text = s(cell(row, 17))
        if s(cell(row, 1)).startswith('COMPOSITE SCORE'):
            composite_score = num(cell(row, 17))

    return {
        'year':            2026,
        'months':          MONTHS,
        'metrics':         metrics,
        'composite_rows':  composite_rows,
        'composite_score': composite_score,
        'rating':          rating_text,
    }


# --- Data Sources ------------------------------------------------------
def extract_sources(sh):
    rows = list(sh.iter_rows(values_only=True))
    items = []
    governance = []
    in_gov = False
    for row in rows:
        kpi_id = s(cell(row, 1))
        if kpi_id == 'GOVERNANCE & CONTROL PRINCIPLES':
            in_gov = True
            continue
        if in_gov:
            label = s(cell(row, 1))
            text  = s(cell(row, 3))
            if label and text:
                governance.append({'label': label, 'text': text})
            continue
        if kpi_id.startswith('KPI-'):
            items.append({
                'id':         kpi_id,
                'kpi':        s(cell(row, 2)),
                'source':     s(cell(row, 3)),
                'system':     s(cell(row, 4)),
                'owner':      s(cell(row, 5)),
                'frequency':  s(cell(row, 6)),
                'retention':  s(cell(row, 7)),
            })
    return {'rows': items, 'governance': governance}


def main():
    if len(sys.argv) < 2:
        print('usage: py extract_hse.py <xlsx-path>', file=sys.stderr)
        sys.exit(2)
    xlsx = sys.argv[1]
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    out = {
        'data_date':       datetime.now().strftime("%Y-%m-%d"),
        'source_file':     xlsx.split('\\')[-1],
        'cover':           extract_cover(wb['Cover']),
        'kpi_definitions': extract_definitions(wb['KPI Definitions']),
        'scoring_matrix':  extract_scoring(wb['Scoring Matrix']),
        'monthly_tracker': extract_tracker(wb['Monthly Tracker']),
        'data_sources':    extract_sources(wb['Data Sources']),
    }
    sys.stdout.reconfigure(encoding='utf-8')
    print(json.dumps(out, ensure_ascii=False, default=str))


if __name__ == '__main__':
    main()
